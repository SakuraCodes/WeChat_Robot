import re

text = "满30.00返15.00 满20.00返10.00"

# 匹配规则(\d+匹配匹配任意数字，\.转义小数点)
pattern = r"满(\d+\.\d+)返(\d+\.\d+)"

matches = re.findall(pattern, text)

for match in matches:
    full_amount, return_amount = match
    print(f"满{full_amount}返{return_amount}")

from openpyxl import load_workbook

# 加载两个Excel文件
workbook1 = load_workbook('book1.xlsx')
workbook2 = load_workbook('book2.xlsx')

# 获取需要操作的工作表
sheet1 = workbook1['Sheet1']
sheet2 = workbook2['Sheet2']

# 假设在sheet1的A列查找值，在sheet2的A列查找，并将sheet2的B列的值写入sheet1的C列
for row in sheet1.iter_rows(min_row=2, values_only=True):  # 从第二行开始，只获取值
    lookup_value = row[0]

    # 在sheet2中查找
    for row2 in sheet2.iter_rows(min_row=2, values_only=True):
        if row2[0] == lookup_value:
            # 找到匹配项，将sheet2的B列值写入sheet1的C列
            sheet1.cell(row=row[1], column=3, value=row2[1])
            break

# 保存结果
workbook1.save('result.xlsx')